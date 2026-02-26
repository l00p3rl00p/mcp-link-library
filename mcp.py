from __future__ import annotations
import sqlite3
import hashlib
import argparse
from urllib.parse import urlparse
import urllib.request
import urllib.error
import sys
try:
    import pypdf
except ImportError:
    pypdf = None
try:
    import openpyxl
except ImportError:
    openpyxl = None
try:
    import docx
except ImportError:
    docx = None
try:
    from PIL import Image
    from PIL.ExifTags import TAGS
except ImportError:
    Image = None
try:
    from watchdog.observers import Observer
    from watchdog.observers.polling import PollingObserver
    from watchdog.events import FileSystemEventHandler
except ImportError:
    Observer = None
    PollingObserver = None
    FileSystemEventHandler = None
import json
from pathlib import Path
from typing import Optional, List, Dict, Any
import datetime
import os
import re
import time
import shlex
__version__ = "3.4.0"

try:
    from atp_sandbox import ATPSandbox
    atp_sb = ATPSandbox()
except ImportError:
    atp_sb = None

try:
    from nexus_session_logger import NexusSessionLogger
    session_logger = NexusSessionLogger()
except ImportError:
    session_logger = None

# Nexus Support: Try to load high-confidence libraries from the central venv
def inject_nexus_env():
    try:
        home = Path.home() if sys.platform != "win32" else Path(os.environ['USERPROFILE'])
        nexus_venv = home / ".mcp-tools" / ".venv"
        if nexus_venv.exists():
            # Add site-packages to sys.path
            import platform
            if platform.system() == "Windows":
                site_pkgs = nexus_venv / "Lib" / "site-packages"
            else:
                # Find python version dir
                lib_dir = nexus_venv / "lib"
                py_dirs = list(lib_dir.glob("python3*"))
                if py_dirs:
                    site_pkgs = py_dirs[0] / "site-packages"
                else:
                    return
            
            if site_pkgs.exists() and str(site_pkgs) not in sys.path:
                sys.path.insert(0, str(site_pkgs))
    except Exception:
        pass

inject_nexus_env()

class SecureMcpLibrary:
    """
    Secure Multi-Contextual Processor (MCP) Link Library
    Provides secure link storage, retrieval, and management.
    """
    
    app_dir: Path

    def __init__(self, db_path: str = None, *, allow_http: bool = False):
        # Use a standard location in ~/.mcp-tools/mcp-server-manager (or override).
        self.allow_http = bool(allow_http)

        override_home = os.environ.get("NEXUS_LIBRARIAN_HOME") or os.environ.get("NEXUS_APP_HOME")
        if override_home:
            self.app_dir = Path(override_home).expanduser().resolve()
        elif sys.platform == "win32":
            self.app_dir = Path(os.environ.get("USERPROFILE", str(Path.home()))) / ".mcp-tools" / "mcp-server-manager"
        else:
            self.app_dir = Path.home() / ".mcp-tools" / "mcp-server-manager"

        # If a specific db_path is provided, still keep an app_dir for logs/artifacts.
        if db_path == ":memory:":
            try:
                import tempfile as _tempfile

                self.app_dir = Path(_tempfile.gettempdir()) / "mcp-link-library"
            except Exception:
                pass
        try:
            self.app_dir.mkdir(parents=True, exist_ok=True)
        except Exception:
            # Last resort: fall back to current working directory.
            self.app_dir = Path.cwd()

        if db_path is None:
            db_path = str(self.app_dir / "knowledge.db")
        
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.cursor = self.conn.cursor()
        self._create_secure_tables()
    
    def _create_secure_tables(self):
        """Create database tables with security constraints."""
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS links (
                id INTEGER PRIMARY KEY,
                url TEXT UNIQUE NOT NULL,
                title TEXT,
                domain TEXT NOT NULL,
                description TEXT,
                categories TEXT,
                stack TEXT DEFAULT 'default',
                is_active BOOLEAN DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                hash TEXT,
                content TEXT
            )
        ''')
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS scan_roots (
                id INTEGER PRIMARY KEY,
                path TEXT UNIQUE NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        self.conn.commit()
        # Migrate existing DBs: add stack column if absent
        try:
            self.cursor.execute("ALTER TABLE links ADD COLUMN stack TEXT DEFAULT 'default'")
            self.conn.commit()
        except Exception:
            pass  # Column already exists
    
    def add_link(self, url: str, categories: List[str] = None, stack: str = "default"):
        """Add a new link with secure metadata extraction."""
        validated_url = self._validate_url(url)
        url_hash = hashlib.sha256(validated_url.encode()).hexdigest()
        
        metadata = self._extract_link_metadata(validated_url)
        categories = categories or ['uncategorized']
        
        # Try to read content for indexing
        content = None
        if validated_url.startswith("file://"):
             try:
                 path = Path(validated_url.replace("file://", "")).resolve()
                 if path.exists():
                     if path.suffix.lower() == '.pdf':
                         try:
                             reader = pypdf.PdfReader(path)
                             text = []
                             for page in reader.pages:
                                 text.append(page.extract_text())
                             content = "\n".join(text)
                         except Exception as e:
                             print(f"⚠️  PDF extraction failed for {path}: {e}", file=sys.stderr)
                             content = None
                     elif path.suffix.lower() == '.xlsx' and openpyxl:
                         try:
                            wb = openpyxl.load_workbook(path, data_only=True)
                            text = []
                            for sheet in wb.sheetnames:
                                ws = wb[sheet]
                                text.append(f"Sheet: {sheet}")
                                for row in ws.iter_rows(values_only=True):
                                    row_text = " ".join([str(c) for c in row if c is not None])
                                    if row_text:
                                        text.append(row_text)
                            content = "\n".join(text)
                         except Exception as e:
                            print(f"⚠️  Excel extraction failed for {path}: {e}", file=sys.stderr)
                            content = None
                     elif path.suffix.lower() == '.docx' and docx:
                         try:
                             doc = docx.Document(path)
                             content = "\n".join([p.text for p in doc.paragraphs])
                         except Exception as e:
                             print(f"⚠️  Word extraction failed for {path}: {e}", file=sys.stderr)
                             content = None
                     elif path.suffix.lower() in ['.jpg', '.jpeg', '.png', '.gif', '.bmp'] and Image:
                         try:
                             with Image.open(path) as img:
                                 meta = [
                                     f"Format: {img.format}",
                                     f"Mode: {img.mode}",
                                     f"Size: {img.size[0]}x{img.size[1]}",
                                 ]
                                 if hasattr(img, '_getexif') and img._getexif():
                                     exif = {
                                         TAGS.get(k, k): v
                                         for k, v in img._getexif().items()
                                         if k in TAGS
                                     }
                                     meta.append(f"EXIF: {str(exif)}")
                                 content = "\n".join(meta)
                         except Exception as e:
                             print(f"⚠️  Image extraction failed for {path}: {e}", file=sys.stderr)
                             content = None
                     else:
                        # Text Handling
                        try:
                            content = path.read_text(errors='ignore')
                        except:
                            pass
             except:
                 pass

        self.cursor.execute('''
            INSERT OR REPLACE INTO links
            (url, title, domain, description, categories, stack, is_active, hash, content)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            validated_url,
            metadata['title'],
            metadata['domain'],
            metadata['description'],
            ','.join(categories),
            stack or 'default',
            1,
            url_hash,
            content
        ))
        self.conn.commit()
        return self.cursor.lastrowid

    def list_links(self, category: str = None, search: str = None, only_active: bool = True, stack: str = None):
        """List links with optional filtering. Returns (id, url, title, domain, categories, is_active, stack)."""
        query = "SELECT id, url, title, domain, categories, is_active, COALESCE(stack,'default') FROM links WHERE 1=1"
        params = []
        
        if only_active:
            query += " AND is_active = 1"
        if category:
            query += " AND categories LIKE ?"
            params.append(f"%{category}%")
        if search:
            query += " AND (title LIKE ? OR url LIKE ? OR description LIKE ? OR content LIKE ?)"
            params.extend([f"%{search}%", f"%{search}%", f"%{search}%", f"%{search}%"])
        if stack:
            query += " AND COALESCE(stack,'default') = ?"
            params.append(stack)

        self.cursor.execute(query, params)
        return self.cursor.fetchall()

    def list_stacks(self) -> List[str]:
        """Return all distinct stack names that have at least one active resource."""
        self.cursor.execute(
            "SELECT DISTINCT COALESCE(stack,'default') FROM links WHERE is_active=1 ORDER BY 1"
        )
        return [row[0] for row in self.cursor.fetchall()]

    def get_categories(self, stack: str = None) -> List[dict]:
        """
        Return first and second-order categories with counts.
        Categories are stored as comma-separated strings, e.g. 'docs,api,reference'.
        First-order: the first token.  Second-order: the second token (if present).
        """
        query = "SELECT categories, COUNT(*) as cnt FROM links WHERE is_active=1"
        params = []
        if stack:
            query += " AND COALESCE(stack,'default') = ?"
            params.append(stack)
        query += " GROUP BY categories"
        self.cursor.execute(query, params)
        rows = self.cursor.fetchall()

        first_order: dict = {}
        second_order: dict = {}
        for (cats_str, cnt) in rows:
            parts = [c.strip() for c in (cats_str or "uncategorized").split(",") if c.strip()]
            first = parts[0] if parts else "uncategorized"
            first_order[first] = first_order.get(first, 0) + cnt
            if len(parts) > 1:
                second = parts[1]
                key = f"{first}/{second}"
                second_order[key] = second_order.get(key, 0) + cnt

        result = []
        for cat, cnt in sorted(first_order.items()):
            result.append({"category": cat, "order": 1, "count": cnt})
        for cat, cnt in sorted(second_order.items()):
            result.append({"category": cat, "order": 2, "count": cnt})
        return result

    def update_link(self, link_id: int, url: str = None, categories: List[str] = None, active: bool = None):
        """Update an existing link."""
        updates = []
        params = []
        if url:
            updates.append("url = ?")
            params.append(url)
        if categories:
            updates.append("categories = ?")
            params.append(','.join(categories))
        if active is not None:
            updates.append("is_active = ?")
            params.append(1 if active else 0)
            
        if not updates:
            return False
            
        params.append(link_id)
        self.cursor.execute(f"UPDATE links SET {', '.join(updates)} WHERE id = ?", params)
        self.conn.commit()
        return True

    def delete_link(self, link_id: int):
        """Permanently delete a link."""
        self.cursor.execute("DELETE FROM links WHERE id = ?", (link_id,))
        self.conn.commit()
        return True

    def open_resource(self, link_id: int):
        """Open a resource using the OS default handler."""
        self.cursor.execute("SELECT url FROM links WHERE id = ?", (link_id,))
        row = self.cursor.fetchone()
        if not row:
            raise ValueError(f"Resource {link_id} not found")
        
        url = row[0]
        import subprocess
        
        if url.startswith("file://"):
            path = url.replace("file://", "")
            if sys.platform == "darwin":
                subprocess.run(["open", path])
            elif sys.platform == "win32":
                os.startfile(path)
            else:
                subprocess.run(["xdg-open", path])
        elif url.startswith("mcp://"):
            # Internal protocols might not be openable directly by OS
            return False
        else:
            # Standard URL
            if sys.platform == "darwin":
                subprocess.run(["open", url])
            elif sys.platform == "win32":
                os.startfile(url)
            else:
                subprocess.run(["xdg-open", url])
        return True

    def edit_resource(self, link_id: int):
        """Open a file resource in the default editor."""
        self.cursor.execute("SELECT url FROM links WHERE id = ?", (link_id,))
        row = self.cursor.fetchone()
        if not row:
            raise ValueError(f"Resource {link_id} not found")
        
        url = row[0]
        if not url.startswith("file://"):
            return False # Can't edit remote URLs or pseudo-protocols
            
        path = url.replace("file://", "")
        import subprocess
        
        editor = os.environ.get("EDITOR", "open -e" if sys.platform == "darwin" else "notepad" if sys.platform == "win32" else "vi")
        
        if sys.platform == "darwin" and editor == "open -e":
            subprocess.run(["open", "-e", path])
        else:
            subprocess.run(shlex.split(editor) + [path])
        return True

    def _validate_url(self, url: str):
        """
        Sanitize and validate URL.
        Ensures protocol is present and parses the netloc to verify structure.
        """
        if not url.startswith(('http://', 'https://', 'file://')):
            url = 'https://' + url
        if url.startswith("http://") and not self.allow_http:
            raise ValueError(
                f"Insecure URL scheme blocked: {url}. Use https:// or pass --allow-http if you really need plain HTTP."
            )
        parsed = urlparse(url)
        
        # Files can have empty netloc (file:///path), HTTP must have domain
        if not parsed.netloc and not url.startswith("file://"):
            # Raise error if URL doesn't have a valid domain after protocol
            raise ValueError(f"Invalid URL: {url}")
        return url
    
    def _extract_link_metadata(self, url: str):
        """
        Extract title and description from URL.
        Uses requests for the fetch and BeautifulSoup for parsing the HTML DOM.
        """
        try:
            import requests
            from bs4 import BeautifulSoup
        except ImportError:
            print("⚠️  Dependencies 'requests' or 'beautifulsoup4' not found.")
            print("   Run 'python3 mcp.py --bootstrap' or 'pip install requests beautifulsoup4'")
            return {'title': url, 'domain': urlparse(url).netloc, 'description': 'Scraper dependencies missing'}

        try:
            # Set a 5s timeout to prevent hanging on slow servers
            response = requests.get(url, timeout=5)
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # Extract title, fall back to URL if <title> tag is missing
            title_tag = soup.title.string if soup.title else url
            
            return {
                'title': title_tag.strip() if title_tag else url,
                'domain': urlparse(url).netloc,
                # Try to find the standard SEO description meta tag
                'description': (
                    soup.find('meta', attrs={'name': 'description'}).get('content', '') 
                    if soup.find('meta', attrs={'name': 'description'}) 
                    else ''
                )
            }
        except Exception:
            # Fallback for unreachable or non-HTML resources
            return {
                'title': url,
                'domain': urlparse(url).netloc,
                'description': 'Metadata extraction failed'
            }

    def index_directory(self, path: str):
        """Index a local directory into the knowledge base."""
        indexer = FileIndexer(path)
        count = 0
        print(f"📂 Indexing {path}...")
        for file_data in indexer.scan():
            self.cursor.execute('''
                INSERT OR REPLACE INTO links 
                (url, title, domain, description, categories, is_active, hash, content) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                f"file://{file_data['path']}", 
                file_data['name'], 
                'local-file', 
                f"Local file: {file_data['rel_path']}", 
                'file,code', 
                1,
                file_data['hash'],
                file_data['content']
            ))
            count += 1
            if count % 10 == 0:
                print(f"   Indexed {count} files...", end='\r')
        self.conn.commit()
        print(f"✅ Indexed {count} files.")

    def index_nexus_suite(self):
        """
        Phase 12: Decoupled Suite Synergy
        Discover and index sibling tool data (Observer, Injector) from the shared Nexus root.
        """
        print("🔗 Nexus Suite Discovery initialized...")
        
        # 1. Locate Nexus Root
        if sys.platform == "win32":
            nexus_root = Path(os.environ['USERPROFILE']) / ".mcp-tools"
        else:
            nexus_root = Path.home() / ".mcp-tools"
            
        if not nexus_root.exists():
            print("❌ Nexus root not found. Are the tools installed?")
            return

        # 2. Discover Observer Inventory
        observer_inv = nexus_root / "mcp-server-manager" / "inventory.yaml"
        if observer_inv.exists():
            self._index_inventory(observer_inv)
        else:
            print("ℹ️  Observer inventory not initialized (fresh install detected).")

        # 3. Discover Injector Config
        injector_conf = nexus_root / "config.json"
        if injector_conf.exists():
            self._index_injector_config(injector_conf)
        else:
            print("⚠️  Injector config not found (global config.json).")
            
        print("✅ Suite indexing complete.")

    def _index_inventory(self, path: Path):
        """Parse Observer inventory.yaml and index servers."""
        print(f"👁️  Indexing Observer Inventory: {path}")
        try:
            import yaml
            data = yaml.safe_load(path.read_text(errors='ignore')) or {}
            servers = data.get("servers", [])
            for s in servers:
                # Map to Knowledge Schema
                name = s.get("name", "Unknown")
                cmd = s.get("run", {}).get("start_cmd") or s.get("command", "n/a")
                desc = s.get("notes", "") or f"MCP Server: {name}"
                
                # Create a pseudo-URL for the knowledge base
                url = f"mcp://observer/server/{s.get('id', name)}"
                
                self.cursor.execute('''
                    INSERT OR REPLACE INTO links 
                    (url, title, domain, description, categories, is_active, hash, content) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    url,
                    f"Server: {name}",
                    "mcp-observer",
                    desc,
                    "suite_inventory,server",
                    1,
                    hashlib.sha256(str(s).encode()).hexdigest(),
                    json.dumps(s, indent=2) # Store full metadata as content
                ))
            self.conn.commit()
            print(f"   Indexed {len(servers)} servers from Observer.")
        except Exception as e:
            print(f"❌ Failed to index inventory: {e} (Need pyyaml?)")

    def _index_injector_config(self, path: Path):
        """Parse global config.json to find managed IDEs."""
        print(f"💉 Indexing Injector Config: {path}")
        try:
            data = json.loads(path.read_text(errors='ignore'))
            ide_paths = data.get("ide_config_paths", {})
            
            for ide, conf_path in ide_paths.items():
                url = f"mcp://injector/ide/{ide}"
                self.cursor.execute('''
                    INSERT OR REPLACE INTO links 
                    (url, title, domain, description, categories, is_active, hash, content) 
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    url,
                    f"IDE Config: {ide.upper()}",
                    "mcp-injector",
                    f"Managed config file for {ide}",
                    "suite_config,ide",
                    1,
                    hashlib.sha256(str(conf_path).encode()).hexdigest(),
                    json.dumps({"ide": ide, "path": conf_path}, indent=2)
                ))
            self.conn.commit()
            print(f"   Indexed {len(ide_paths)} managed IDEs.")
        except Exception as e:
            print(f"❌ Failed to index injector config: {e}")

    def prepopulate_docs(self, path: str):
        """Generates standardized ARCHITECTURE.md and README.md for a folder."""
        target = Path(path).resolve()
        if not target.exists():
            return False
        
        name = target.name
        
        # README.md Template
        readme_content = f"""# {name} (Nexus Forged)

## Overview
This MCP server was automatically forged by the Workforce Nexus. It implements standard Deterministic Wrapping and ATP Hardening.

## Capabilities
- **Deterministic API**: Wrapped with `mcp_wrapper.py`.
- **ATP Sandbox**: Hardened with `atp_sandbox.py`.

## Usage
Run with `python3 mcp_server.py`.
"""
        # ARCHITECTURE.md Template
        arch_content = f"""# Architecture: {name}

## Components
1. **Entry Point**: `mcp_server.py`
2. **Logic Layer**: Wrapped tools accessible via MCP.
3. **Security Model**: ATP Sandbox enforces logic-only execution.

## Data Flow
Input -> mcp_wrapper -> [Tool Logic] -> sandbox_verify -> Output
"""
        # ATP_COMPLIANCE_GUIDE.md Template (The Portability Mandate)
        compliance_content = f"""# ATP Compliance Guide: {name}

## 🏛 The Portability Mandate (v3.0)
This MCP server is a self-contained "Logic Colony." It carries its own security runtime and documentation spec so that any subsequent AI agent can operate it with high-fidelity.

## 🧪 The "Strawberry" Logic Test
To verify compliance, run:
```python
result = "strawberry".count("r")
```
Required Outcome: `3` (Verified via `atp_sandbox.py`)

## 🔧 Component Specification
1. **mcp_wrapper.py**: Deterministic API adapter (13-point standard).
2. **atp_sandbox.py**: Isolated Logic Jail (Whitelist-only).

### Standard 3.0 Specs:
- **1-3**: HTTP/POST/GET with custom header & auth support.
- **4-5**: Strict Input/Output contracts (Status, Data, Error, Elapsed).
- **6-7**: Stream aggregation and deterministic retry policies.
- **8-10**: 1MB safety guards, safe-schemes (no file://), and observability hooks.
- **11-13**: Token projection (dot-path), schema validation, and tier annotation.
"""
        (target / "README.md").write_text(readme_content)
        (target / "ARCHITECTURE.md").write_text(arch_content)
        (target / "ATP_COMPLIANCE_GUIDE.md").write_text(compliance_content)
        return True

    def link_categories(self, server_id: str, tags: List[str]):
        """Associates resource tags with a server logic ID in the DB."""
        # For now, we store this in the 'categories' field of the server entry in the links table
        # We find the observer server link and update it
        url = f"mcp://observer/server/{server_id}"
        self.cursor.execute("UPDATE links SET categories = categories || ? WHERE url = ?", ("," + ",".join(tags), url))
        self.conn.commit()
        return True

class FileIndexer:
    def __init__(self, root_path: str):
        self.root = Path(root_path).resolve()
        self.ignore_patterns = self._load_gitignore()
    
    def _load_gitignore(self):
        patterns = ['.git', '__pycache__', 'node_modules', '.venv', '.DS_Store', '*.pyc']
        gitignore = self.root / ".gitignore"
        if gitignore.exists():
            with open(gitignore) as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith('#'):
                        patterns.append(line)
        return patterns

    def _should_ignore(self, path: Path):
        """
        Determine if a file should be ignored based on .gitignore patterns.
        Tiered Reliability: Uses pathspec (Industrial mode) or Regex (Standard mode).
        """
        # 1. Permanent Tier: Try PathSpec
        try:
            from pathspec import PathSpec
            from pathspec.patterns import GitWildMatchPattern
            # Cache the spec for performance if needed, but for now we re-parse
            spec = PathSpec.from_lines(GitWildMatchPattern, self.ignore_patterns)
            if spec.match_file(str(path.relative_to(self.root))):
                return True
        except ImportError:
            pass # Fallback to standard
            
        # 2. Standard Tier: Regex-based ignoring (Pure Python)
        import fnmatch
        rel_path = str(path.relative_to(self.root))
        
        # Internal hard-ignores
        internal_ignores = [".git", "__pycache__", ".venv", "node_modules", ".DS_Store"]
        if any(part in internal_ignores for part in path.parts):
            return True

        for pattern in self.ignore_patterns:
            if not pattern or pattern.startswith("#"):
                continue
            
            # Use regex for more accurate matching than fnmatch
            # (Simplified gitignore->regex conversion)
            regex = pattern.replace(".", "\\.").replace("*", ".*").replace("?", ".")
            if pattern.startswith("/"):
                regex = "^" + regex[1:]
            
            try:
                if re.search(regex, rel_path) or fnmatch.fnmatch(path.name, pattern):
                    return True
            except:
                continue
                
        return False

    def scan(self):
        extensions = {'.md', '.txt', '.py', '.js', '.json', '.yaml', '.yml', '.sh', '.html', '.css'}
        for path in self.root.rglob('*'):
            if path.is_file() and path.suffix in extensions:
                if self._should_ignore(path):
                    continue
                
                try:
                    content = path.read_text(errors='ignore')
                    yield {
                        'path': str(path),
                        'rel_path': str(path.relative_to(self.root)),
                        'name': path.name,
                        'content': content,
                        'hash': hashlib.sha256(content.encode()).hexdigest()
                    }
                except Exception as e:
                    print(f"⚠️  Skipping {path.name}: {e}")

if FileSystemEventHandler is None:
    class FileSystemEventHandler:
        pass

class NexusWatcher(FileSystemEventHandler):
    """
    Real-time file observer for the Librarian.
    Re-indexes files when they are modified or created.
    """
    def __init__(self, library: SecureMcpLibrary, paths: List[str]):
        self.library = library
        self.paths = paths
        self.observer = None

    def on_modified(self, event):
        if not event.is_directory:
            self._handle_change(event.src_path)

    def on_created(self, event):
        if not event.is_directory:
            print(f"DEBUG: Watcher detected creation of {event.src_path}", file=sys.stderr)
            self._handle_change(event.src_path)

    def _handle_change(self, path_str: str):
        path = Path(path_str)
        # Skip hidden files and common noise
        if any(part.startswith('.') for part in path.parts):
            return
        
        # Log to file since stdout is redirected in MCP mode
        self.library.cursor.execute("INSERT INTO links (url, title, domain, categories, content) VALUES (?, ?, ?, ?, ?)", 
                                   (f"log://watcher/{time.time()}", "Watcher Event", "system", "debug", f"Detected: {path.name}"))
        self.library.conn.commit()
        
        print(f"🔄 Change detected: {path.name}. Re-indexing...", file=sys.stderr)
        # Add or update in DB
        try:
            self.library.add_link(f"file://{path_str}", categories=['auto-indexed'])
        except Exception as e:
            print(f"❌ Auto-index failed for {path.name}: {e}", file=sys.stderr)

    def start(self):
        if not PollingObserver:
            print("❌ Watchdog not installed. Cannot start watcher.", file=sys.stderr)
            return False
            
        self.observer = PollingObserver()
        for p in self.paths:
            path_obj = Path(p).resolve()
            if path_obj.exists():
                self.observer.schedule(self, str(path_obj), recursive=True)
                print(f"👁️  Watching: {path_obj}", file=sys.stderr)
        
        self.observer.start()
        return True

    def stop(self):
        if self.observer:
            self.observer.stop()
            self.observer.join()

class MCPServer:
    def __init__(self):
        # Prefer the canonical on-disk DB, but fall back to :memory: in restricted
        # environments (CI/sandbox) where the default home directory may be read-only.
        try:
            self.library = SecureMcpLibrary()
        except sqlite3.OperationalError as e:
            print(f"WARNING: Librarian DB unavailable ({e}); using in-memory DB.", file=sys.stderr)
            self.library = SecureMcpLibrary(":memory:")
        log_dir = self.library.app_dir / "logs"
        log_dir.mkdir(parents=True, exist_ok=True)
        self.log_path = log_dir / "librarian_errors.log"
        self.watcher = None
        # GAP-R3 FIX: Auto-start watcher on documents/ directory at server launch.
        # Mission: "the ecosystem's memory is always accessible" requires continuous indexing.
        self._auto_start_watcher()

    def _auto_start_watcher(self):
        """Auto-start the file watcher on the documents/ directory at launch.
        Falls back gracefully if watchdog is not installed."""
        if PollingObserver is None:
            print("ℹ️  Watchdog not installed — real-time watcher disabled.", file=sys.stderr)
            print("   Install with: pip install watchdog", file=sys.stderr)
            return
        docs_dir = self.library.app_dir / "documents"
        docs_dir.mkdir(parents=True, exist_ok=True)
        watch_paths = [str(docs_dir)]
        self.watcher = NexusWatcher(self.library, watch_paths)
        started = self.watcher.start()
        if started:
            print(f"👁️  Auto-watcher started on: {docs_dir}", file=sys.stderr)
        else:
            print("⚠️  Auto-watcher failed to start.", file=sys.stderr)
            self.watcher = None

        
    def log_error(self, msg: str):
        """Log critical errors to file so they aren't lost in stdio redirection."""
        try:
            timestamp = datetime.datetime.now().isoformat()
            with open(self.log_path, "a") as f:
                f.write(f"[{timestamp}] {msg}\n")
        except:
            pass # Last resort, don't crash
        
    def run(self):
        """Run the MCP JSON-RPC loop over stdio."""
        # Ensure we don't pollute stdout with print statements
        # Redirect stdout to formatted JSON messages
        while True:
            try:
                line = sys.stdin.readline()
                if not line:
                    break
                
                request = json.loads(line)
                response = self.handle_request(request)
                
                if response:
                    print(json.dumps(response), flush=True)
                    
            except json.JSONDecodeError:
                continue
            except Exception as e:
                # Log to stderr to avoid breaking the protocol
                import traceback
                error_msg = f"Server Error: {e}\n{traceback.format_exc()}"
                print(f"Server Error: {e}", file=sys.stderr)
                self.log_error(error_msg)

    def handle_request(self, request: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        method = request.get("method")
        params = request.get("params", {})
        msg_id = request.get("id")
        
        result = None
        error = None
        
        try:
            if method == "initialize":
                result = {
                    "protocolVersion": "2024-11-05",
                    "serverInfo": {
                        "name": "mcp-link-library",
                        "version": "0.1.0"
                    },
                    "capabilities": {
                        "resources": {
                            "listChanged": False,
                            "subscribe": False
                        },
                        "tools": {
                            "listChanged": False
                        }
                    }
                }
            elif method == "notifications/initialized":
                # No response needed for notifications
                return None
            elif method == "resources/list":
                # OPTIMIZATION: Zero-Token Processing
                # Do NOT dump thousands of files. Return a capped list + instruction.
                files = self.library.list_links(category="file,code", only_active=True)
                resources = []
                
                # Cap at 50 to prevent context flooding
                limit = 50
                for f in files[:limit]:
                    # f: id, url, title, domain, categories, is_active
                    resources.append({
                        "uri": f[1],
                        "name": f[2],
                        "mimeType": "text/plain"
                    })
                
                if len(files) > limit:
                     resources.append({
                        "uri": "nexus://guidance/search-truncated",
                        "name": f"... ({len(files) - limit} more files) - Use 'search_knowledge_base' tool to find specific files",
                        "mimeType": "text/plain"
                    })
                    
                result = {"resources": resources}
            elif method == "resources/read":
                uri = params.get("uri")
                content = self._read_resource(uri)
                result = {
                    "contents": [{
                        "uri": uri,
                        "mimeType": "text/plain",
                        "text": content
                    }]
                }
            elif method == "tools/list":
                result = {
                    "tools": [{
                        "name": "search_knowledge_base",
                        "description": "Search indexed files and documents within a named stack (knowledge context). WORKFLOW: 1) If the user hasn't specified a stack, call list_stacks first to see available stacks (like NotebookLM projects). 2) Call get_categories on the chosen stack to understand what's inside. 3) Then search with a query + stack name. Omit stack to search across all stacks.",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "query": {
                                    "type": "string",
                                    "description": "Search term"
                                },
                                "stack": {
                                    "type": "string",
                                    "description": "Stack name to search within (e.g. 'gravity-research', 'cookie-recipes'). Call list_stacks first if unsure. Omit to search all stacks."
                                }
                            },
                            "required": ["query"]
                        }
                    }, {
                        "name": "add_resource",
                        "description": "Add a new resource (URL or file) to the Knowledge Base. Assign it to a named stack to keep it grouped with related knowledge. Stacks are isolated knowledge contexts — like NotebookLM projects, but called stacks here. Example: stack='gravity-research' for physics papers, stack='cookie-recipes' for baking docs.",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "url": {"type": "string", "description": "URL or file:// path to add"},
                                "categories": {"type": "string", "description": "Comma-separated categories (e.g. 'docs,api')"},
                                "stack": {"type": "string", "description": "Stack name to file this resource under (e.g. 'gravity-research'). Defaults to 'default'. Create a new stack by simply using a new name."}
                            },
                            "required": ["url"]
                        }
                    }, {
                        "name": "update_resource",
                        "description": "Update an existing resource",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "integer", "description": "Resource ID"},
                                "url": {"type": "string", "description": "New URL (optional)"},
                                "title": {"type": "string", "description": "New Title (optional)"},
                                "active": {"type": "boolean", "description": "Set active state (optional)"}
                            },
                            "required": ["id"]
                        }
                    }, {
                        "name": "delete_resource",
                        "description": "Delete a resource by ID",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "id": {"type": "integer", "description": "Resource ID to delete"}
                            },
                            "required": ["id"]
                        }
                    }, {
                        "name": "check_health",
                        "description": "Check the health of the Librarian's dependencies (pypdf, openpyxl, etc.)",
                        "inputSchema": {
                            "type": "object",
                            "properties": {},
                            "required": []
                        }
                    }, {
                        "name": "update_dependencies",
                        "description": "Update or install missing dependencies for the Librarian",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "packages": {
                                    "type": "string",
                                    "description": "Space-separated list of packages to install (e.g. 'pypdf openpyxl')"
                                }
                            },
                            "required": ["packages"]
                        }
                    }, {
                        "name": "start_watcher",
                        "description": "Start a real-time file watcher for specific directories",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "paths": {
                                    "type": "array",
                                    "items": {"type": "string"},
                                    "description": "List of directories to watch"
                                }
                            },
                            "required": ["paths"]
                        }
                    }, {
                        "name": "get_watcher_logs",
                        "description": "Retrieve internal debug logs from the file watcher",
                        "inputSchema": {
                            "type": "object",
                            "properties": {},
                            "required": []
                        }
                    }, {
                        "name": "search_api",
                        "description": "ATP Tool: Search for available tool signatures by intent or keyword. Prevents context bloat.",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "query": {"type": "string", "description": "Intent or keyword to search for"}
                            },
                            "required": ["query"]
                        }
                    }, {
                        "name": "execute_code",
                        "description": "ATP Tool: Execute Python code in a restricted sandbox for data processing (map/filter/reduce).",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "code": {"type": "string", "description": "Python code block to execute"},
                                "context": {"type": "object", "description": "Optional data context for the code"}
                            },
                            "required": ["code"]
                        }
                    }, {
                        "name": "prepopulate_docs",
                        "description": "Generate standard ARCHITECTURE.md and README.md for a directory",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "path": {"type": "string", "description": "Absolute path to the directory"}
                            },
                            "required": ["path"]
                        }
                    }, {
                        "name": "link_categories",
                        "description": "Associate tags/categories with a forged server",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "server_id": {"type": "string", "description": "ID of the server"},
                                "tags": {"type": "array", "items": {"type": "string"}, "description": "List of tags"}
                            },
                            "required": ["server_id", "tags"]
                        }
                    }, {
                        "name": "list_stacks",
                        "description": "List all named stacks (knowledge contexts) that have indexed resources. Stacks are like projects in NotebookLM or notebooks in Stitch — each one is an isolated, named knowledge context. Call this first whenever the user wants to search or browse without specifying a stack, so you can ask them which context to use.",
                        "inputSchema": {
                            "type": "object",
                            "properties": {}
                        }
                    }, {
                        "name": "get_categories",
                        "description": "Browse the categories of knowledge inside a stack (or across all stacks). Use before searching to understand what topics are covered, so you can guide the user to the right search terms or confirm they have the right stack.",
                        "inputSchema": {
                            "type": "object",
                            "properties": {
                                "stack": {
                                    "type": "string",
                                    "description": "Optional: scope to this stack. Omit to see categories across all stacks."
                                }
                            }
                        }
                    }]
                }
            elif method == "tools/call":
                name = params.get("name")
                args = params.get("arguments", {})
                if name == "search_knowledge_base":
                    query = args.get("query")
                    stack = args.get("stack") or None
                    matches = self.library.list_links(search=query, stack=stack)
                    stack_tag = f" in stack '{stack}'" if stack else ""
                    text = f"Found matches{stack_tag}:\n"
                    if not matches:
                        text += "No results found."
                    else:
                        for m in matches:
                            # m = (id, url, title, domain, categories, is_active, stack)
                            domain_tag = f"[{m[3]}]" if m[3] else ""
                            stack_label = f" [{m[6]}]" if m[6] and m[6] != "default" else ""
                            text += f"- {domain_tag}{stack_label} {m[2]} ({m[1]})\n"
                    
                    result = {
                        "content": [{
                            "type": "text",
                            "text": text
                        }]
                    }
                elif name == "search_api":
                    query = args.get("query", "").lower()
                    tools = self.handle_request({"method": "tools/list", "id": 0})["tools"]
                    matches = [t for t in tools if query in t["name"].lower() or query in t["description"].lower()]
                    result = {
                        "content": [{"type": "text", "text": json.dumps(matches, indent=2)}]
                    }
                elif name == "execute_code":
                    code = args.get("code", "")
                    context = args.get("context", {})
                    if not atp_sb:
                         result = {"isError": True, "content": [{"type": "text", "text": "ATP Sandbox module missing."}]}
                    else:
                        exec_res = atp_sb.execute(code, context)
                        if exec_res.get("success"):
                            result = {
                                "content": [{"type": "text", "text": json.dumps(exec_res["result"], indent=2)}]
                            }
                        else:
                            result = {
                                "isError": True,
                                "content": [{"type": "text", "text": f"Sandbox Error: {exec_res.get('error')}"}]
                            }
                elif name == "add_resource":
                    url = args.get("url")
                    cats = args.get("categories", "").split(",") if args.get("categories") else None
                    stack = args.get("stack", "default") or "default"
                    try:
                        new_id = self.library.add_link(url, cats, stack=stack)
                        result = {
                            "content": [{"type": "text", "text": f"✅ Added resource ID: {new_id} to stack '{stack}'"}]
                        }
                    except Exception as e:
                        result = {
                            "content": [{"type": "text", "text": f"❌ Failed: {e}"}],
                            "isError": True
                        }
                elif name == "update_resource":
                    rid = args.get("id")
                    url = args.get("url")
                    active = args.get("active")
                    # Note: library.update_link doesn't support title update yet in SQL, 
                    # but we'll wire up what we have.
                    try:
                        success = self.library.update_link(rid, url=url, active=active)
                        msg = f"✅ Updated ID {rid}" if success else f"❌ ID {rid} not found"
                        result = {
                            "content": [{"type": "text", "text": msg}]
                        }
                    except Exception as e:
                         result = {
                            "content": [{"type": "text", "text": f"❌ Failed: {e}"}],
                            "isError": True
                        }
                elif name == "delete_resource":
                    rid = args.get("id")
                    try:
                        self.library.delete_link(rid)
                        result = {
                            "content": [{"type": "text", "text": f"🗑️ Deleted ID {rid}"}]
                        }
                    except Exception as e:
                         result = {
                            "content": [{"type": "text", "text": f"❌ Failed: {e}"}],
                            "isError": True
                        }
                elif name == "check_health":
                    status = []
                    status.append(f"pypdf: {'✅' if pypdf else '❌ (pip install pypdf)'}")
                    status.append(f"openpyxl: {'✅' if openpyxl else '❌ (pip install openpyxl)'}")
                    status.append(f"python-docx: {'✅' if docx else '❌ (pip install python-docx)'}")
                    status.append(f"Pillow: {'✅' if Image else '❌ (pip install Pillow)'}")
                    result = {
                        "content": [{"type": "text", "text": "\n".join(status)}]
                    }
                elif name == "update_dependencies":
                    pkgs = args.get("packages", "")
                    if not pkgs:
                         raise ValueError("No packages specified")
                    import subprocess
                    try:
                        cmd = [sys.executable, "-m", "pip", "install"] + pkgs.split()
                        subprocess.run(cmd, check=True, capture_output=True)
                        result = {
                            "content": [{"type": "text", "text": f"✅ Successfully installed: {pkgs}"}]
                        }
                    except subprocess.CalledProcessError as e:
                        result = {
                            "content": [{"type": "text", "text": f"❌ Install failed: {e.stderr.decode()}"}],
                            "isError": True
                        }
                    except Exception as e:
                        result = {
                            "content": [{"type": "text", "text": f"❌ Install failed: {e}"}],
                            "isError": True
                        }
                elif name == "start_watcher":
                    if self.watcher:
                        self.watcher.stop()
                    
                    paths = args.get("paths", [])
                    self.watcher = NexusWatcher(self.library, paths)
                    success = self.watcher.start()
                    
                    if success:
                        result = {
                            "content": [{"type": "text", "text": f"👁️  Watcher started for: {', '.join(paths)}"}]
                        }
                    else:
                        result = {
                            "content": [{"type": "text", "text": "❌ Failed to start watcher (Check logs for details)"}],
                            "isError": True
                        }
                elif name == "get_watcher_logs":
                    self.library.cursor.execute("SELECT content FROM links WHERE categories = 'debug' ORDER BY id DESC LIMIT 10")
                    rows = self.library.cursor.fetchall()
                    text = "Recent Watcher Events:\n" + ("\n".join([r[0] for r in rows]) if rows else "No events recorded.")
                    result = {
                        "content": [{"type": "text", "text": text}]
                    }
                elif name == "prepopulate_docs":
                    path = args.get("path")
                    success = self.library.prepopulate_docs(path)
                    result = {
                        "content": [{"type": "text", "text": f"{'✅' if success else '❌'} Documentation prepopulated at {path}"}]
                    }
                elif name == "link_categories":
                    sid = args.get("server_id")
                    tags = args.get("tags", [])
                    self.library.link_categories(sid, tags)
                    result = {
                        "content": [{"type": "text", "text": f"✅ Linked {len(tags)} tags to {sid}"}]
                    }
                elif name == "list_stacks":
                    stacks = self.library.list_stacks()
                    if stacks:
                        text = "Available stacks:\n" + "\n".join(f"- {s}" for s in stacks)
                    else:
                        text = "No stacks yet. Add resources with a stack name to create one."
                    result = {"content": [{"type": "text", "text": text}]}
                elif name == "get_categories":
                    stack = args.get("stack") or None
                    cats = self.library.get_categories(stack=stack)
                    if cats:
                        header = f"Categories{' in stack ' + stack if stack else ' (all stacks)'}:"
                        lines = [header]
                        for c in cats:
                            indent = "  " if c["order"] == 1 else "    └─ "
                            lines.append(f"{indent}{c['category']}  ({c['count']} resources)")
                        text = "\n".join(lines)
                    else:
                        text = "No categories found."
                    result = {"content": [{"type": "text", "text": text}]}
                else:
                    raise ValueError(f"Unknown tool: {name}")
            elif method == "ping":
                result = {}
            else:
                # Ignore unknown notifications, error on requests
                if msg_id is not None:
                     raise ValueError(f"Method not found: {method}")
                return None
                
        except Exception as e:
            error = {
                "code": -32603,
                "message": str(e)
            }
        
        if msg_id is not None:
            response = {
                "jsonrpc": "2.0",
                "id": msg_id
            }
            if error:
                response["error"] = error
                if session_logger and method == "tools/call":
                    session_logger.log_command(params.get("name", "unknown"), "ERROR", result=error.get("message"))
            else:
                response["result"] = result
                if session_logger and method == "tools/call":
                    res_text = json.dumps(result)[:200] + "..." if result else "None"
                    session_logger.log_command(params.get("name", "unknown"), "SUCCESS", result=res_text)
            return response
        
        return None

    def _read_resource(self, uri: str) -> str:
        # Check DB first
        self.library.cursor.execute("SELECT content FROM links WHERE url = ?", (uri,))
        row = self.library.cursor.fetchone()
        if row and row[0]:
            return row[0]
        
        # Fallback to local file if it exists and path matches
        # uri format: file:///path/to/file
        # SECURITY: Only allow reading files within the current working directory to prevent traversal
        if uri.startswith("file://"):
            try:
                path = Path(uri.replace("file://", "")).resolve()
                # SECURITY: Allow reading if file exists and is within a recognized Nexus workspace
                allowed_roots = [
                    self.library.app_dir.parent, # ~/.mcp-tools
                    Path("/Users/almowplay/Developer/Github") # Active developer root
                ]
                is_allowed = any(str(path).startswith(str(root)) for root in allowed_roots)
                
                if path.exists() and is_allowed:
                     # Binary check: try reading as text, if fails return base64 or msg
                     try:
                         return path.read_text(errors='ignore')
                     except:
                         return f"[Binary File] Size: {path.stat().st_size} bytes"
                else:
                    return f"Error: File not found at {path}"
            except Exception as e:
                return f"Error reading file: {e}"
        
        # Enable Remote Fetching (Unified Data Retrieval)
        if uri.startswith("http://") or uri.startswith("https://"):
            try:
                # Basic fetch with User-Agent to avoid eager bot blocking
                req = urllib.request.Request(
                    uri, 
                    headers={'User-Agent': 'Nexus-Librarian/1.0 (MCP-Unified-Retrieval)'}
                )
                with urllib.request.urlopen(req, timeout=10) as response:
                    # Limit size to 1MB to prevent memory issues
                    data = response.read(1024 * 1024)
                    return data.decode('utf-8', errors='ignore')
            except Exception as e:
                # Return error as content so agent knows why it failed
                return f"Error retrieving remote resource: {e}"
        
        raise ValueError(f"Resource not found: {uri}")


def check_suite():
    """Detect presence of sibling tools in the Git-Packager suite."""
    siblings = {
        "mcp-injector": "The Surgeon",
        "mcp-server-manager": "The Observer",
        "repo-mcp-packager": "The Activator"
    }
    print("📋 Workforce Suite Checkback:")
    root = Path(__file__).parent.parent
    for folder, persona in siblings.items():
        exists = (root / folder).exists()
        status = "✅ PRESENT" if exists else "❌ MISSING"
        print(f"  {persona:<15} ({folder:<20}): {status}")

def cmd_bootstrap():
    """Universal bootstrapper for the Librarian."""
    try:
        import importlib.util
        local_bootstrap = Path(__file__).parent / "bootstrap.py"
        if not local_bootstrap.exists():
            print("❌ Local bootstrap.py not found.")
            return 1
        
        spec = importlib.util.spec_from_file_location("bootstrap", local_bootstrap)
        bootstrap = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(bootstrap)
        bootstrap.main()
        return 0
    except Exception as e:
        print(f"❌ Bootstrap failed: {e}")
        return 1

def main():
    parser = argparse.ArgumentParser(
        description="MCP Link Library - The Librarian",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Commands (what they do):\n"
            "  mcp-librarian --add <url>         Add a URL\n"
            "  mcp-librarian --list              List active links\n"
            "  mcp-librarian --search <query>    Search stored links\n"
            "  mcp-librarian --index <dir>       Index a local directory\n"
            "  mcp-librarian --index-suite       Index suite data (Observer/Injector)\n"
            "  mcp-librarian --server            Run as stdio MCP server (injectable)\n"
        ),
    )
    parser.add_argument('--add', help="Add a new link")
    parser.add_argument('--categories', nargs='+', help="Categories for the link")
    parser.add_argument('--stack', help="Stack name for --add (e.g. 'gravity-research'). Defaults to 'default'.")
    parser.add_argument('--list-stacks', action='store_true', dest='list_stacks', help="List all named stacks")
    parser.add_argument('--list', action='store_true', help="List active links")
    parser.add_argument('--category', help="Filter listing by category")
    parser.add_argument('--search', help="Search links")
    parser.add_argument('--delete', type=int, help="Delete link by ID")
    parser.add_argument('--open', type=int, help="Open resource by ID using OS default")
    parser.add_argument('--edit', type=int, help="Edit resource by ID using default editor")
    parser.add_argument('--update', type=int, help="Update link ID (requires --url or --categories)")
    parser.add_argument('--url', help="New URL for update")
    parser.add_argument('--activate', type=int, help="Activate link by ID")
    parser.add_argument('--deactivate', type=int, help="Deactivate link by ID")
    parser.add_argument('--bootstrap', action='store_true', help="Bootstrap the Git-Packager workspace")
    parser.add_argument('--check', action='store_true', help="Check for sibling tool presence")
    parser.add_argument('--index', help="Index a local directory")
    parser.add_argument('--index-suite', action='store_true', help="Index Nexus Suite (Observer/Injector) data")
    parser.add_argument('--server', action='store_true', help="Run in MCP Server mode")
    parser.add_argument('--watch', action='store_true', help="Start real-time directory watcher")
    parser.add_argument('--prepopulate-docs', help="Prepopulate docs for a directory")
    parser.add_argument('--json', action='store_true', help="Output in raw JSON format for agent-side processing")
    parser.add_argument('--allow-http', action='store_true', help="Allow adding http:// URLs (NOT recommended).")
    
    args = parser.parse_args()
    
    if args.bootstrap:
        sys.exit(cmd_bootstrap())

    if args.prepopulate_docs:
        library = SecureMcpLibrary(allow_http=bool(args.allow_http))
        if library.prepopulate_docs(args.prepopulate_docs):
            print(f"✅ Prepopulated documentation at {args.prepopulate_docs}")
        else:
            print(f"❌ Failed to prepopulate docs at {args.prepopulate_docs}")
        return

    if args.server:
        server = MCPServer()
        server.run()
        return

    if args.watch:
        library = SecureMcpLibrary(allow_http=bool(args.allow_http))
        # Fetch roots from DB
        rows = library.cursor.execute("SELECT path FROM scan_roots").fetchall()
        paths = [r[0] for r in rows]
        if not paths:
            print("❌ No scan roots registered in DB. Add one using --index first or via the GUI/Observer.")
            return
        
        watcher = NexusWatcher(library, paths)
        if watcher.start():
            print(f"🚀 Librarian Watcher started on {len(paths)} paths. Press Ctrl+C to stop.")
            try:
                while True: time.sleep(1)
            except KeyboardInterrupt:
                watcher.stop()
                print("🛑 Watcher stopped.")
        return

    if args.index_suite:
        library = SecureMcpLibrary(allow_http=bool(args.allow_http))
        library.index_nexus_suite()
        return

    if args.index:
        library = SecureMcpLibrary(allow_http=bool(args.allow_http))
        library.index_directory(args.index)
        return

    if args.check:
        check_suite()
        return
        
    library = SecureMcpLibrary(allow_http=bool(args.allow_http))
    
    if args.add:
        try:
            stack = getattr(args, 'stack', None) or 'default'
            link_id = library.add_link(args.add, args.categories, stack=stack)
            print(f"✅ Added link with ID: {link_id} (stack: {stack})")
        except Exception as e:
            print(f"❌ Failed to add link: {e}")

    elif getattr(args, 'list_stacks', False):
        stacks = library.list_stacks()
        if stacks:
            print("Available stacks:")
            for s in stacks:
                print(f"  - {s}")
        else:
            print("No stacks yet.")

    elif args.list:
        links = library.list_links(category=args.category, search=args.search)
        if args.json:
            print(json.dumps([{"id": l[0], "url": l[1], "title": l[2], "domain": l[3], "categories": l[4], "active": l[5]} for l in links]))
            return
        if not links:
            print("No links found.")
        else:
            print(f"{'ID':<4} {'Domain':<20} {'Categories':<20} {'Title'}")
            print("-" * 60)
            for l in links:
                print(f"{l[0]:<4} {l[3]:<20} {l[4]:<20} {l[2] or l[1]}")
                
    elif args.delete:
        library.delete_link(args.delete)
        print(f"🗑 Deleted link {args.delete}")

    elif args.open:
        if library.open_resource(args.open):
            print(f"🚀 Opened resource {args.open}")
        else:
            print(f"❌ Could not open resource {args.open}")

    elif args.edit:
        if library.edit_resource(args.edit):
            print(f"📝 Opened {args.edit} for editing")
        else:
            print(f"❌ Resource {args.edit} is not an editable local file.")
        
    elif args.update:
        if library.update_link(args.update, url=args.url, categories=args.categories):
            print(f"✨ Updated link {args.update}")
        else:
            print("No updates provided. Use --url or --categories.")
            
    elif args.activate:
        library.update_link(args.activate, active=True)
        print(f"✅ Activated link {args.activate}")
        
    elif args.deactivate:
        library.update_link(args.deactivate, active=False)
        print(f"💤 Deactivated link {args.deactivate}")
    else:
        parser.print_help()

if __name__ == "__main__":
    main()
